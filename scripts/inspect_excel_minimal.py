
import openpyxl
import sys

file_path = "EXCEL_CONSOLIDATO_FINALE.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    print("Sheet names:", wb.sheetnames)
    
    for sheet in wb.sheetnames:
        print(f"\nScanning Sheet: {sheet}")
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            row_str = str(row)
            # Check for rank related info
            if "3.85" in row_str or "3.85" in row_str or "4.19" in row_str:
                print(f"FOUND ROW: {row}")

except Exception as e:
    print(f"Error: {e}")
